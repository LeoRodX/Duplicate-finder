import os
import time
from datetime import datetime
import dbf
import xlrd
from openpyxl import load_workbook
import humanize
import sys
import chardet

def detect_encoding(file_path):
    """Определить кодировку файла"""
    try:
        with open(file_path, 'rb') as f:
            rawdata = f.read(10000)
            result = chardet.detect(rawdata)
            return result['encoding'] or 'unknown'
    except:
        return 'unknown'

def get_files_in_directory(path):
    """Получить список XLS, XLSX и DBF файлов в указанной директории"""
    files = []
    for file in os.listdir(path):
        if file.lower().endswith(('.xls', '.xlsx', '.dbf')):
            files.append(file)
    return files

def select_file(files):
    """Выбрать файл из списка"""
    print("\nДоступные файлы:")
    for i, file in enumerate(files, 1):
        print(f"{i}. {file}")
    
    while True:
        try:
            choice = int(input("\nВведите номер файла: "))
            if 1 <= choice <= len(files):
                return files[choice - 1]
            print("Некорректный номер. Попробуйте снова.")
        except ValueError:
            print("Пожалуйста, введите число.")

def get_columns(file_path):
    """Получить список столбцов из файла и дополнительную информацию"""
    encoding = detect_encoding(file_path)
    
    if file_path.lower().endswith('.dbf'):
        try:
            table = dbf.Table(file_path)
            table.open()
            
            if hasattr(table, 'field_names'):
                columns = table.field_names
            elif hasattr(table, 'fields'):
                columns = [field.name for field in table.fields]
            else:
                columns = []
                
            row_count = len(table)
            
            try:
                if hasattr(table, 'codepage'):
                    from dbf.code_pages import CODEPAGES
                    encoding = CODEPAGES.get(table.codepage, encoding)
            except:
                pass
                
            table.close()
            return columns, row_count, encoding
            
        except Exception as e:
            print(f"Ошибка при чтении DBF файла {file_path}: {str(e)}")
            return None, None, encoding
    
    elif file_path.lower().endswith('.xlsx'):
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            columns = [cell.value for cell in sheet[1]] if sheet[1] else []
            row_count = sheet.max_row  # Учитываем заголовок
            wb.close()
            return columns, row_count, 'utf-8'
        except Exception as e:
            print(f"Ошибка при чтении XLSX файла: {e}")
            return None, None, encoding
    
    else:  # старый .xls
        try:
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            columns = sheet.row_values(0)
            row_count = sheet.nrows  # Учитываем заголовок
            return columns, row_count, 'cp1251'
        except Exception as e:
            print(f"Ошибка при чтении XLS файла: {e}")
            return None, None, encoding

def select_column(columns):
    """Выбрать столбец из списка"""
    print("\nДоступные столбцы:")
    for i, column in enumerate(columns, 1):
        print(f"{i}. {column}")
    
    while True:
        try:
            choice = int(input("\nВведите номер столбца: "))
            if 1 <= choice <= len(columns):
                return columns[choice - 1], choice - 1
            print("Некорректный номер. Попробуйте снова.")
        except ValueError:
            print("Пожалуйста, введите число.")

def find_duplicates(file_path, column_index):
    """Найти дубликаты в указанном столбце"""
    duplicates = {}
    
    if file_path.lower().endswith('.dbf'):
        try:
            table = dbf.Table(file_path)
            table.open()
            
            if hasattr(table, 'field_names'):
                field_name = table.field_names[column_index]
                # Для DBF добавляем +1 к номеру строки, чтобы соответствовать LibreOffice
                for i, record in enumerate(table, 2):  # Начинаем с 2, как в LibreOffice
                    value = record[field_name]
                    if value in duplicates:
                        duplicates[value].append(i)
                    else:
                        duplicates[value] = [i]
            else:
                for i, record in enumerate(table, 2):  # Начинаем с 2, как в LibreOffice
                    value = record[column_index]
                    if value in duplicates:
                        duplicates[value].append(i)
                    else:
                        duplicates[value] = [i]
                        
            table.close()
        except Exception as e:
            print(f"Ошибка при поиске дубликатов в DBF: {e}")
    
    elif file_path.lower().endswith('.xlsx'):
        try:
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            # Нумерация строк соответствует Excel (первая строка данных = 2)
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if column_index < len(row):
                    value = row[column_index]
                    if value in duplicates:
                        duplicates[value].append(i)
                    else:
                        duplicates[value] = [i]
            wb.close()
        except Exception as e:
            print(f"Ошибка при поиске дубликатов в XLSX: {e}")
    
    else:  # .xls
        try:
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            # Нумерация строк соответствует Excel (первая строка данных = 2)
            for i in range(1, sheet.nrows):
                value = sheet.cell_value(i, column_index)
                if value in duplicates:
                    duplicates[value].append(i + 1)  # +1 для соответствия Excel
                else:
                    duplicates[value] = [i + 1]
        except Exception as e:
            print(f"Ошибка при поиске дубликатов в XLS: {e}")
    
    return {k: v for k, v in duplicates.items() if len(v) > 1}

def save_report(report_path, selected_file, file_path, file_size, encoding, row_count, search_time, column_name, duplicates):
    """Сохранить отчет о дубликатах"""
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("="*50 + "\n")
        f.write("ОТЧЕТ О ПОИСКЕ ДУБЛИКАТОВ\n")
        f.write("="*50 + "\n\n")
        
        f.write(f"Имя файла: {selected_file}\n")
        f.write(f"Путь к файлу: {file_path}\n")
        f.write(f"Размер файла: {humanize.naturalsize(file_size)}\n")
        f.write(f"Кодировка файла: {encoding or 'не определена'}\n")
        f.write(f"Количество строк в таблице (с заголовком): {row_count}\n")
        f.write(f"Длительность поиска: {search_time:.2f} сек.\n")
        f.write(f"Анализируемый столбец: {column_name}\n\n")
        
        f.write("="*50 + "\n")
        f.write("НАЙДЕННЫЕ ДУБЛИКАТЫ\n")
        f.write("="*50 + "\n\n")
        
        if duplicates:
            f.write(f"Всего найдено дублирующихся значений: {len(duplicates)}\n\n")
            for i, (value, rows) in enumerate(duplicates.items(), 1):
                f.write(f"{i}. Значение: {repr(value)}\n")
                f.write(f"   Строки: {', '.join(map(str, sorted(rows)))}\n\n")
        else:
            f.write("Дубликаты не найдены.\n")

def main():
    print("Поиск дубликатов в XLS, XLSX и DBF файлах\n")
    
    # Запрашиваем путь к папке
    while True:
        path = input("Введите путь к папке с файлами: ").strip()
        if os.path.isdir(path):
            break
        print("Указанная папка не существует. Попробуйте снова.")
    
    # Получаем список файлов
    files = get_files_in_directory(path)
    if not files:
        print("В указанной папке нет XLS, XLSX или DBF файлов.")
        input("Нажмите Enter для выхода...")
        return
    
    # Выбираем файл
    selected_file = select_file(files)
    file_path = os.path.join(path, selected_file)
    print(f"\nВыбран файл: {file_path}")
    
    # Получаем информацию о файле
    file_size = os.path.getsize(file_path)
    
    # Получаем столбцы и дополнительную информацию
    columns, row_count, encoding = get_columns(file_path)
    if not columns:
        input("Нажмите Enter для выхода...")
        return
    
    # Выбираем столбец
    column_name, column_index = select_column(columns)
    
    # Ищем дубликаты
    print("\nПоиск дубликатов...")
    start_time = time.time()
    duplicates = find_duplicates(file_path, column_index)
    search_time = time.time() - start_time
    
    # Выводим результаты
    if duplicates:
        print(f"\nНайдено {len(duplicates)} дублирующихся значений:")
        for i, (value, rows) in enumerate(duplicates.items(), 1):
            print(f"{i}. Значение: {value}")
            print(f"   Строки: {', '.join(map(str, sorted(rows)))}\n")
    else:
        print("\nДубликаты не найдены.")
    
    # Сохраняем отчет
    reports_dir = os.path.join(os.path.dirname(__file__), "Отчеты")
    os.makedirs(reports_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(selected_file)[0]
    report_filename = f"report_{base_name}_{timestamp}.txt"
    report_path = os.path.join(reports_dir, report_filename)
    
    save_report(report_path, selected_file, file_path, file_size, encoding, 
               row_count, search_time, column_name, duplicates)
    
    print(f"\nОтчет сохранен в: {report_path}")
    print(f"\nПоиск занял {search_time:.2f} секунд.")
    
    # Ожидаем ввода перед закрытием
    input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Произошла непредвиденная ошибка: {e}")
        import traceback
        traceback.print_exc()
        input("Нажмите Enter для выхода...")